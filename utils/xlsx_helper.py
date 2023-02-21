import os

from openpyxl import Workbook, load_workbook

from utils.table import Table


class XlsxHelper(object):

    def __init__(self, file_name='demo.xlsx'):
        self.file_name = file_name
        if os.path.exists(self.file_name):
            self.wb = load_workbook(filename=file_name)
        else:
            self.wb = Workbook()

    def __getitem__(self, item):
        if item in self.wb:
            return self.wb.__getitem__(item)
        else:
            self.wb.create_sheet(title=item)
            return self.__getitem__(item)

    def save(self):
        self.wb.save(filename=self.file_name)


class EvalDataSaver(object):

    def __init__(self, file_name='/home/mi-nan/workspace/myPyCharm/elsp_drl/experiment/eval_result.xlsx'):
        self.xlsx_helper = XlsxHelper(file_name=file_name)
        self.model_path = None
        self.algorithm = None
        self.env_id = None
        self.seed = None
        self.already_row = 0
        self.__computed__ = {}
        self.col = 0
        self.row = 0

    def __getattr__(self, key):
        if key in self.__computed__:
            return self.__computed__[key]
        else:
            return self.__dict__[key]

    def get_col_str(self, col=None):
        if col is None:
            col = self.col
        col_str = ''
        if col >= 26:
            col_str += self.get_col_str(col//26) + self.get_col_str(col % 26)
        else:
            col_str += str(bytes([b'A'[0] + col]), encoding='utf-8')
        return col_str

    def __setattr__(self, key, value):
        if '__computed__' in self.__dict__:
            if key in self.__computed__:
                self.__computed__[key] = value
            else:
                self.__dict__[key] = value
                if key == 'algorithm' or key == 'env_id' or key == 'seed':
                    self.__computed__['sheet_name'] =\
                        ('{}'.format(self.algorithm)) +\
                        ('' if self.env_id is None else ('_env_{}'.format(self.env_id))) +\
                        ('' if self.seed is None else ('_seed_{}'.format(self.seed)))
                elif key == 'row' or key == 'col':
                    try:
                        row_str = str(self.row + 1)
                    except KeyError:
                        row_str = '1'
                    self.__computed__['cell_str'] = '{}{}'.format(self.get_col_str(), row_str)
        else:
            self.__dict__[key] = value

    def set_cell(self, row, col, val):
        self.row = row
        self.col = col
        self.xlsx_helper[self.sheet_name][self.cell_str] = val

    def get_already_row_num(self):
        row_buff = self.row
        col_buff = self.col
        self.col = 0
        self.row = 0
        ws = self.xlsx_helper[self.sheet_name]
        while ws[self.cell_str].value is not None:
            self.row += 1
        row = self.row
        self.row = row_buff
        self.col = col_buff
        return row - 1

    def save_eval_data(self, model_path, eval_data, row_offset=0, col_offset=0):
        ws = self.xlsx_helper[self.sheet_name]
        if model_path is not None:
            self.col = 1 + col_offset
            self.row = 0 + row_offset
            cell1 = self.cell_str
            self.col = 4 + col_offset
            cell2 = self.cell_str
            ws.merge_cells('{}:{}'.format(cell1, cell2))
            self.col = 0 + col_offset
            ws[self.cell_str] = model_path
            ws[cell1] = os.path.basename(model_path)
        for r in range(len(eval_data)):
            self.row = r + 1 + row_offset
            for c in range(len(eval_data[r])):
                self.col = c + col_offset
                cell_data = eval_data[r][c]
                if cell_data is not None:
                    ws[self.cell_str] = cell_data

    def append_eval_data(self, model_path, eval_data):
        self.save_eval_data(model_path, eval_data, self.get_already_row_num())

    def set_eval_data(self, model_path, demand_scale, eval_data):
        self.col = 0
        self.row = 0
        ws = self.xlsx_helper[self.sheet_name]
        while ws[self.cell_str].value is not None:
            if ws[self.cell_str].value == model_path:
                break
            self.row += 1
        row_offset = self.row
        self.row += 1
        while ws[self.cell_str].value is not None:
            if ws[self.cell_str].value == demand_scale:
                break
            self.col += 1
        col_offset = self.col
        self.save_eval_data(model_path, eval_data, row_offset, col_offset)

    def load_eval_data(self, col, row, w, h):
        eval_data = []
        ws = self.xlsx_helper[self.sheet_name]
        for i in range(row, row + h):
            self.row = i
            eval_data.append([])
            for j in range(col, col + w):
                self.col = j
                eval_data[-1].append(ws[self.cell_str].value)
        return eval_data

    def load_row_data_list(self, row, w, h):
        self.row = row
        col = 0
        self.col = 0
        ws = self.xlsx_helper[self.sheet_name]
        data_list = []
        while ws[self.cell_str].value is not None:
            data_list.append(self.load_eval_data(col, row, w, h))
            col += w
            self.col = col
            self.row = row
        return data_list

    def sort(self, row, w, h):
        eval_data_list = self.load_row_data_list(row, w, h)
        eval_data_list.sort()
        for i in range(len(eval_data_list)):
            eval_data = eval_data_list[i]
            self.save_eval_data(None, eval_data, row_offset=row-1, col_offset=i*w)


class DataAnalyser(object):

    def __init__(self, file_name):
        self.file_name = file_name
        self.eval_data_saver1 = EvalDataSaver(file_name=file_name)
        self.eval_data_saver2 = EvalDataSaver(file_name=file_name)

    def analyse_all_on_demand(self, algorithm1, algorithm2, _seed, row, w, h, percent=False):
        self.eval_data_saver1.algorithm = algorithm1
        self.eval_data_saver1.seed = _seed
        self.eval_data_saver2.algorithm = algorithm2
        self.eval_data_saver2.seed = _seed
        row_data_list_1 = self.eval_data_saver1.load_row_data_list(row, w, h)
        row_data_1 = {}
        for i in range(len(row_data_list_1)):
            row_data = row_data_list_1[i]
            demand_scale = row_data[0][0]
            row_data[0][0] = 'name'
            head_list = row_data.pop(0)
            table = Table(row_data, head_list, row_head_key='name')
            row_data_1[demand_scale] = table
        row_data_list_2 = self.eval_data_saver2.load_row_data_list(row, w, h)
        row_data_2 = {}
        for i in range(len(row_data_list_2)):
            row_data = row_data_list_2[i]
            demand_scale = row_data[0][0]
            row_data[0][0] = 'name'
            head_list = row_data.pop(0)
            table = Table(row_data, head_list, row_head_key='name')
            row_data_2[demand_scale] = table

        res_saver = EvalDataSaver(file_name=self.file_name)
        res_saver.algorithm = 'cmp{}_{}_{}'.format('_percent' if percent else '', algorithm1, algorithm2)
        res_saver.seed = _seed
        for demand_scale in row_data_1:
            data1 = row_data_1[demand_scale]
            data2 = row_data_2[demand_scale]
            res = [data1.table_head_list]
            res[0][0] = demand_scale
            for row_name in data1.data_dict(deep_copy=False):
                res.append([])
                for col_name in data1.data_dict(deep_copy=False)[row_name]:
                    if col_name == 'name':
                        res[-1].append(data2[row_name][col_name])
                    else:
                        res[-1].append((data2[row_name][col_name] - data1[row_name][col_name]) / (data1[row_name][col_name] if percent else 1))
            res_saver.save_eval_data(None, res, row, list(row_data_1.keys()).index(demand_scale) * w)
        res_saver.xlsx_helper.save()

    def analyse_on_demand(self, row_name, col_name, algorithm1, algorithm2, _seed, row, w, h, percent=False):
        self.eval_data_saver1.algorithm = algorithm1
        self.eval_data_saver1.seed = _seed
        self.eval_data_saver2.algorithm = algorithm2
        self.eval_data_saver2.seed = _seed
        row_data_list_1 = self.eval_data_saver1.load_row_data_list(row, w, h)
        row_data_1 = {}
        for i in range(len(row_data_list_1)):
            row_data = row_data_list_1[i]
            demand_scale = row_data[0][0]
            row_data[0][0] = 'name'
            head_list = row_data.pop(0)
            table = Table(row_data, head_list, row_head_key='name')
            row_data_1[demand_scale] = table
        row_data_list_2 = self.eval_data_saver2.load_row_data_list(row, w, h)
        row_data_2 = {}
        for i in range(len(row_data_list_2)):
            row_data = row_data_list_2[i]
            demand_scale = row_data[0][0]
            row_data[0][0] = 'name'
            head_list = row_data.pop(0)
            table = Table(row_data, head_list, row_head_key='name')
            row_data_2[demand_scale] = table
        res_saver = EvalDataSaver(file_name=self.file_name)
        res_saver.algorithm = 'cmp_{}_{}{}'.format(row_name, col_name, '_percent' if percent else '')
        res_saver.seed = _seed
        res = []
        for demand_scale in row_data_1:
            data1 = row_data_1[demand_scale]
            data2 = row_data_2[demand_scale]
            res.append((data2[row_name][col_name] - data1[row_name][col_name]) / (data1[row_name][col_name] if percent else 1))
        res_saver.save_eval_data(None, [res], res_saver.get_already_row_num(), 0)
        res_saver.xlsx_helper.save()


if __name__ == '__main__':
    # _model_path = '/home/mi-nan/workspace/myPyCharm/elsp_drl/experiment/standard_self_attention/result/env_path_elsp_env_00105/algorithm_PPO/total_1.100000e+07_n_steps_2.000000e+02_batch_size_6.400000e+03_n_epochs_5.000000e+01_env_num_32_lr_1.000000e-03_gama_0.960/time_2021--03--08 20-43-57/ef_195306.24_mean_p_194118.99_max_p198800.42.zip'
    # env_id = 5
    # seed = 32
    # eval_data_saver = EvalDataSaver()
    # eval_data_saver.algorithm = 'PPO_SSA'
    # eval_data_saver.env_id = '10{}'.format(env_id)
    # eval_data_saver.seed = seed
    # eval_data_saver.sort(1, 5, 5)
    # buff = [
    #     ['统计结果', 'mean', 'std', 'min', 'max'],
    #     ['Profit'],
    #     ['Storage'],
    #     ['CSL'],
    #     ['Backlog penalties'],
    # ]
    # buff = [[2.5, 'mean', 'std', 'min', 'max'], ['Profit', 337771.4445578475, 7089.060687501504, 346928.33049573307, 320470.5484620647], ['Storage', 2046.6237788857275, 173.94130121103714, 2373.2033036270745, 1552.2626415546301], ['CSL', 73.34881147397125, 0.8720045388083101, 74.72266591552705, 71.32172728662667], ['Backlog penalties', 45936.55370334917, 2381.7875678222545, 52146.33645298281, 41990.32794111519]]

    # print(eval_data_saver.get_already_row_num())
    # eval_data_saver.save_eval_data(_model_path, buff)
    # eval_data_saver.set_eval_data(_model_path, 2.5,  buff)
    #
    # def set_cell(col, row, val):
    #     row_str = str(row + row_num * already_num + 1)
    #     col_str = str(bytes([b'A'[0] + col]), encoding='utf-8')
    #     xlsx_helper['PPO_SSA_env_10{}_seed_{}'.format(env_id, seed)]['{}{}'.format(col_str, row_str)] = val
    #
    # def save_eval_data(model_path, eval_data):
    #     set_cell(0, 0, model_path)
    #     .merge_cells('A2:D2')
    #     for r in range(len(eval_data)):
    #         row_data = eval_data[r]
    #         for c in range(len(row_data)):
    #             cell_data = row_data[c]
    #             if cell_data is not None:
    #                 set_cell(c, r + 1, cell_data)
    #
    # save_eval_data(_model_path, buff)
    # eval_data_saver.xlsx_helper.save()
    data_analyser = DataAnalyser('../res/cmp_cma-es_ssa2ppo/demand.xlsx')
    w = 5
    h = 7
    y = 0
    seed = 32
    for i in range(4):
        h += 1
        y += 0 if i == 0 else (h - 1)
        print(h, y)
        data_analyser.analyse_all_on_demand('CMA-ES+BSP2', 'PPO_SSA_F', seed, y, w, h)
        data_analyser.analyse_all_on_demand('CMA-ES+BSP2', 'PPO_SSA_F', seed, y, w, h, True)
        data_analyser.analyse_on_demand('Profit', 'mean', 'CMA-ES+BSP2', 'PPO_SSA_F', seed, y, w, h)
        data_analyser.analyse_on_demand('Profit', 'mean', 'CMA-ES+BSP2', 'PPO_SSA_F', seed, y, w, h, True)
    pass
